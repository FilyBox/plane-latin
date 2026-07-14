# Copyright (c) 2023-present Plane Software, Inc. and contributors
# SPDX-License-Identifier: AGPL-3.0-only
# See the LICENSE file for details.

"""Workspace-level budgets and the expense ledger.

Admin-only: every handler here is gated to ROLE.ADMIN. Members and guests get a
403, and the sidebar entry is hidden from them.

Internal bookkeeping only — nothing here moves money. Budgets allocate an amount
to a category for a period; expenses record what was actually spent, with the
invoice kept as a library FileAsset. "Spent" is always aggregated from the
ledger, never stored, so it cannot drift away from the rows it summarizes.
"""

# Python imports
import csv
from io import BytesIO, StringIO
from decimal import Decimal

# Django imports
from django.db.models import Count, Q, Sum
from django.http import HttpResponse
from django.utils.text import slugify

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

# Third party imports
from rest_framework import status
from rest_framework.response import Response

# Module imports
from plane.app.permissions import ROLE, allow_permission
from plane.app.serializers import (
    BudgetBonusSerializer,
    BudgetScenarioEmployeeSerializer,
    BudgetScenarioSerializer,
    BudgetScenarioVariableSerializer,
    BudgetSerializer,
    ExpenseCategorySerializer,
    ExpenseSerializer,
    FinancialVariableSerializer,
)
from plane.db.models import (
    Budget,
    BudgetBonus,
    BudgetCellOverride,
    BudgetScenario,
    BudgetScenarioEmployee,
    BudgetScenarioVariable,
    Expense,
    ExpenseCategory,
    ExpenseDocument,
    FileAsset,
    FinancialVariable,
    Workspace,
    WorkspaceFeature,
)
from plane.settings.storage import S3Storage
from plane.utils.budgeting import scenario_forecast
from plane.utils.csv_utils import sanitize_csv_row
from plane.utils.workspace_feature import is_workspace_feature_enabled

from ..base import BaseAPIView

CENTS = Decimal("0.01")


def _scenario_forecast_data(scenario):
    employees = (
        BudgetScenarioEmployee.objects.filter(scenario=scenario)
        .select_related("employee", "salary__office")
        .prefetch_related("bonuses")
    )
    variables = BudgetScenarioVariable.objects.filter(scenario=scenario).select_related(
        "variable__office"
    )
    expenses = Expense.objects.filter(
        workspace_id=scenario.workspace_id,
        expense_date__gte=scenario.period_start,
        expense_date__lte=scenario.period_end,
    ).exclude(status=Expense.Status.CANCELLED).select_related("category")
    overrides = BudgetCellOverride.objects.filter(scenario=scenario)
    return scenario_forecast(scenario, employees, variables, expenses, overrides)


class FinanceBaseView(BaseAPIView):
    """Base view enforcing the per-workspace payments feature flag."""

    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        slug = kwargs.get("slug")
        if slug and not is_workspace_feature_enabled(WorkspaceFeature.FeatureKey.PAYMENTS, slug=slug):
            self.permission_denied(request, message="Payments are not enabled for this workspace")


class ExpenseCategoryEndpoint(FinanceBaseView):
    serializer_class = ExpenseCategorySerializer
    model = ExpenseCategory

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        categories = ExpenseCategory.objects.filter(workspace__slug=slug).annotate(
            expense_count=Count("expenses", filter=Q(expenses__deleted_at__isnull=True))
        )
        return Response(ExpenseCategorySerializer(categories, many=True).data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        workspace = Workspace.objects.get(slug=slug)
        serializer = ExpenseCategorySerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        if ExpenseCategory.objects.filter(workspace=workspace, name__iexact=serializer.validated_data["name"]).exists():
            return Response(
                {"name": ["A category with this name already exists"]}, status=status.HTTP_409_CONFLICT
            )
        serializer.save(workspace_id=workspace.id)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ExpenseCategoryDetailEndpoint(FinanceBaseView):
    serializer_class = ExpenseCategorySerializer
    model = ExpenseCategory

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, category_id):
        category = ExpenseCategory.objects.get(id=category_id, workspace__slug=slug)
        serializer = ExpenseCategorySerializer(category, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, category_id):
        category = ExpenseCategory.objects.get(id=category_id, workspace__slug=slug)
        # Expenses keep their history (category goes NULL); budgets are the
        # allocation for a bucket that no longer exists, so they go with it.
        category.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BudgetEndpoint(FinanceBaseView):
    serializer_class = BudgetSerializer
    model = Budget

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        budgets = Budget.objects.filter(workspace__slug=slug)
        category_id = request.query_params.get("category")
        if category_id:
            budgets = budgets.filter(category_id=category_id)
        project_id = request.query_params.get("project")
        if project_id == "none":
            budgets = budgets.filter(project__isnull=True)
        elif project_id:
            budgets = budgets.filter(project_id=project_id)
        # Budgets overlapping the requested window, not only those inside it
        active_on = request.query_params.get("active_on")
        if active_on:
            budgets = budgets.filter(period_start__lte=active_on, period_end__gte=active_on)
        return Response(BudgetSerializer(budgets, many=True).data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        workspace = Workspace.objects.get(slug=slug)
        serializer = BudgetSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        # Surface the "already budgeted" case as a 409 instead of letting the
        # unique constraint raise a 500
        duplicate = Budget.objects.filter(
            workspace=workspace,
            category=serializer.validated_data["category"],
            project=serializer.validated_data.get("project"),
            period_start=serializer.validated_data["period_start"],
            period_end=serializer.validated_data["period_end"],
        ).exists()
        if duplicate:
            return Response(
                {"error": "This category already has a budget for that period"},
                status=status.HTTP_409_CONFLICT,
            )

        serializer.save(workspace_id=workspace.id)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class BudgetDetailEndpoint(FinanceBaseView):
    serializer_class = BudgetSerializer
    model = Budget

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, budget_id):
        budget = Budget.objects.get(id=budget_id, workspace__slug=slug)
        serializer = BudgetSerializer(budget, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, budget_id):
        Budget.objects.get(id=budget_id, workspace__slug=slug).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BudgetScenarioEndpoint(FinanceBaseView):
    serializer_class = BudgetScenarioSerializer
    model = BudgetScenario

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        scenarios = BudgetScenario.objects.filter(workspace__slug=slug).annotate(
            employee_count=Count("employees", filter=Q(employees__deleted_at__isnull=True), distinct=True),
            variable_count=Count("variables", filter=Q(variables__deleted_at__isnull=True), distinct=True),
        )
        year = request.query_params.get("year")
        if year:
            scenarios = scenarios.filter(fiscal_year=year)
        return Response(BudgetScenarioSerializer(scenarios, many=True).data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        workspace = Workspace.objects.get(slug=slug)
        serializer = BudgetScenarioSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        if BudgetScenario.objects.filter(
            workspace=workspace,
            name__iexact=serializer.validated_data["name"],
            fiscal_year=serializer.validated_data["fiscal_year"],
        ).exists():
            return Response(
                {"name": ["A budget with this name already exists for that year"]},
                status=status.HTTP_409_CONFLICT,
            )
        serializer.save(workspace_id=workspace.id)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class BudgetScenarioDetailEndpoint(FinanceBaseView):
    serializer_class = BudgetScenarioSerializer
    model = BudgetScenario

    def _scenario(self, slug, scenario_id):
        return BudgetScenario.objects.filter(id=scenario_id, workspace__slug=slug).annotate(
            employee_count=Count("employees", filter=Q(employees__deleted_at__isnull=True), distinct=True),
            variable_count=Count("variables", filter=Q(variables__deleted_at__isnull=True), distinct=True),
        ).get()

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug, scenario_id):
        return Response(BudgetScenarioSerializer(self._scenario(slug, scenario_id)).data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, scenario_id):
        scenario = self._scenario(slug, scenario_id)
        serializer = BudgetScenarioSerializer(scenario, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        name = serializer.validated_data.get("name", scenario.name)
        fiscal_year = serializer.validated_data.get("fiscal_year", scenario.fiscal_year)
        if BudgetScenario.objects.filter(
            workspace_id=scenario.workspace_id, name__iexact=name, fiscal_year=fiscal_year
        ).exclude(id=scenario.id).exists():
            return Response(
                {"name": ["A budget with this name already exists for that year"]},
                status=status.HTTP_409_CONFLICT,
            )
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, scenario_id):
        self._scenario(slug, scenario_id).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class FinancialVariableEndpoint(FinanceBaseView):
    serializer_class = FinancialVariableSerializer
    model = FinancialVariable

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        variables = FinancialVariable.objects.filter(workspace__slug=slug).select_related("office")
        office_id = request.query_params.get("office")
        if office_id:
            variables = variables.filter(office_id=office_id)
        return Response(FinancialVariableSerializer(variables, many=True).data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        workspace = Workspace.objects.get(slug=slug)
        serializer = FinancialVariableSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        office = serializer.validated_data["office"]
        if office.workspace_id != workspace.id:
            return Response({"office": ["The entity does not belong to this workspace"]}, status=status.HTTP_400_BAD_REQUEST)
        if FinancialVariable.objects.filter(
            workspace=workspace, office=office, name__iexact=serializer.validated_data["name"]
        ).exists():
            return Response(
                {"name": ["A variable with this name already exists for that entity"]},
                status=status.HTTP_409_CONFLICT,
            )
        serializer.save(workspace_id=workspace.id)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class FinancialVariableDetailEndpoint(FinanceBaseView):
    serializer_class = FinancialVariableSerializer
    model = FinancialVariable

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, variable_id):
        variable = FinancialVariable.objects.get(id=variable_id, workspace__slug=slug)
        serializer = FinancialVariableSerializer(variable, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        office = serializer.validated_data.get("office", variable.office)
        if office.workspace_id != variable.workspace_id:
            return Response({"office": ["The entity does not belong to this workspace"]}, status=status.HTTP_400_BAD_REQUEST)
        name = serializer.validated_data.get("name", variable.name)
        if FinancialVariable.objects.filter(
            workspace_id=variable.workspace_id, office=office, name__iexact=name
        ).exclude(id=variable.id).exists():
            return Response(
                {"name": ["A variable with this name already exists for that entity"]},
                status=status.HTTP_409_CONFLICT,
            )
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, variable_id):
        FinancialVariable.objects.get(id=variable_id, workspace__slug=slug).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BudgetScenarioEmployeeEndpoint(FinanceBaseView):
    serializer_class = BudgetScenarioEmployeeSerializer
    model = BudgetScenarioEmployee

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug, scenario_id):
        assignments = BudgetScenarioEmployee.objects.filter(
            scenario_id=scenario_id, workspace__slug=slug
        ).select_related("employee", "salary__office").prefetch_related("bonuses")
        return Response(BudgetScenarioEmployeeSerializer(assignments, many=True).data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug, scenario_id):
        scenario = BudgetScenario.objects.get(id=scenario_id, workspace__slug=slug)
        serializer = BudgetScenarioEmployeeSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        employee = serializer.validated_data["employee"]
        salary = serializer.validated_data["salary"]
        if employee.workspace_id != scenario.workspace_id or salary.workspace_id != scenario.workspace_id:
            return Response({"error": "Employee and salary must belong to this workspace"}, status=status.HTTP_400_BAD_REQUEST)
        start = serializer.validated_data["effective_from"]
        end = serializer.validated_data.get("effective_to") or scenario.period_end
        if start < scenario.period_start or end > scenario.period_end:
            return Response(
                {"error": "The employee dates must stay inside the budget period"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        if BudgetScenarioEmployee.objects.filter(
            scenario=scenario, employee=employee, salary=salary
        ).exists():
            return Response(
                {"error": "This salary is already included for the employee"},
                status=status.HTTP_409_CONFLICT,
            )
        serializer.save(workspace_id=scenario.workspace_id, scenario_id=scenario.id)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class BudgetScenarioEmployeeDetailEndpoint(FinanceBaseView):
    serializer_class = BudgetScenarioEmployeeSerializer
    model = BudgetScenarioEmployee

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, scenario_id, assignment_id):
        assignment = BudgetScenarioEmployee.objects.get(
            id=assignment_id, scenario_id=scenario_id, workspace__slug=slug
        )
        serializer = BudgetScenarioEmployeeSerializer(assignment, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        employee = serializer.validated_data.get("employee", assignment.employee)
        salary = serializer.validated_data.get("salary", assignment.salary)
        if employee.workspace_id != assignment.workspace_id or salary.workspace_id != assignment.workspace_id:
            return Response({"error": "Employee and salary must belong to this workspace"}, status=status.HTTP_400_BAD_REQUEST)
        start = serializer.validated_data.get("effective_from", assignment.effective_from)
        end = serializer.validated_data.get("effective_to", assignment.effective_to) or assignment.scenario.period_end
        if start < assignment.scenario.period_start or end > assignment.scenario.period_end:
            return Response(
                {"error": "The employee dates must stay inside the budget period"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, scenario_id, assignment_id):
        BudgetScenarioEmployee.objects.get(
            id=assignment_id, scenario_id=scenario_id, workspace__slug=slug
        ).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BudgetBonusEndpoint(FinanceBaseView):
    serializer_class = BudgetBonusSerializer
    model = BudgetBonus

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug, scenario_id, assignment_id):
        assignment = BudgetScenarioEmployee.objects.get(
            id=assignment_id, scenario_id=scenario_id, workspace__slug=slug
        )
        serializer = BudgetBonusSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        start = serializer.validated_data["effective_from"]
        end = serializer.validated_data.get("effective_to") or assignment.effective_to or assignment.scenario.period_end
        assignment_end = assignment.effective_to or assignment.scenario.period_end
        if start < assignment.effective_from or end > assignment_end:
            return Response(
                {"error": "The bonus dates must stay inside the employee assignment"},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer.save(workspace_id=assignment.workspace_id, scenario_employee_id=assignment.id)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class BudgetBonusDetailEndpoint(FinanceBaseView):
    serializer_class = BudgetBonusSerializer
    model = BudgetBonus

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, scenario_id, assignment_id, bonus_id):
        bonus = BudgetBonus.objects.get(
            id=bonus_id,
            scenario_employee_id=assignment_id,
            scenario_employee__scenario_id=scenario_id,
            workspace__slug=slug,
        )
        serializer = BudgetBonusSerializer(bonus, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, scenario_id, assignment_id, bonus_id):
        BudgetBonus.objects.get(
            id=bonus_id,
            scenario_employee_id=assignment_id,
            scenario_employee__scenario_id=scenario_id,
            workspace__slug=slug,
        ).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BudgetScenarioVariableEndpoint(FinanceBaseView):
    serializer_class = BudgetScenarioVariableSerializer
    model = BudgetScenarioVariable

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug, scenario_id):
        assignments = BudgetScenarioVariable.objects.filter(
            scenario_id=scenario_id, workspace__slug=slug
        ).select_related("variable__office")
        return Response(BudgetScenarioVariableSerializer(assignments, many=True).data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug, scenario_id):
        scenario = BudgetScenario.objects.get(id=scenario_id, workspace__slug=slug)
        serializer = BudgetScenarioVariableSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        variable = serializer.validated_data["variable"]
        if variable.workspace_id != scenario.workspace_id:
            return Response({"variable": ["The variable does not belong to this workspace"]}, status=status.HTTP_400_BAD_REQUEST)
        if BudgetScenarioVariable.objects.filter(scenario=scenario, variable=variable).exists():
            return Response({"error": "This variable is already included"}, status=status.HTTP_409_CONFLICT)
        serializer.save(workspace_id=scenario.workspace_id, scenario_id=scenario.id)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class BudgetScenarioVariableDetailEndpoint(FinanceBaseView):
    serializer_class = BudgetScenarioVariableSerializer
    model = BudgetScenarioVariable

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, scenario_id, assignment_id):
        BudgetScenarioVariable.objects.get(
            id=assignment_id, scenario_id=scenario_id, workspace__slug=slug
        ).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BudgetScenarioSummaryEndpoint(FinanceBaseView):
    model = BudgetScenario

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug, scenario_id):
        scenario = BudgetScenario.objects.get(id=scenario_id, workspace__slug=slug)
        return Response(_scenario_forecast_data(scenario), status=status.HTTP_200_OK)


class BudgetScenarioCellOverrideEndpoint(FinanceBaseView):
    model = BudgetCellOverride

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def put(self, request, slug, scenario_id):
        scenario = BudgetScenario.objects.get(id=scenario_id, workspace__slug=slug)
        row_key = str(request.data.get("row_key", "")).strip()
        try:
            year = int(request.data.get("year"))
            month = int(request.data.get("month"))
            amount = Decimal(str(request.data.get("amount")))
        except (TypeError, ValueError, ArithmeticError):
            return Response({"error": "Enter a valid amount and month"}, status=status.HTTP_400_BAD_REQUEST)

        forecast = _scenario_forecast_data(scenario)
        valid_rows = {line["key"] for line in forecast["lines"]}
        valid_months = {(item["year"], item["month"]) for item in forecast["months"]}
        if row_key not in valid_rows or (year, month) not in valid_months:
            return Response({"error": "This budget cell does not exist"}, status=status.HTTP_400_BAD_REQUEST)
        if amount < 0:
            return Response({"amount": ["The amount cannot be negative"]}, status=status.HTTP_400_BAD_REQUEST)

        override, _ = BudgetCellOverride.objects.update_or_create(
            scenario=scenario,
            row_key=row_key,
            year=year,
            month=month,
            defaults={"amount": amount, "workspace_id": scenario.workspace_id},
        )
        return Response(
            {
                "id": str(override.id),
                "row_key": override.row_key,
                "year": override.year,
                "month": override.month,
                "amount": str(override.amount),
            },
            status=status.HTTP_200_OK,
        )

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, scenario_id):
        scenario = BudgetScenario.objects.get(id=scenario_id, workspace__slug=slug)
        BudgetCellOverride.objects.filter(
            scenario=scenario,
            row_key=request.data.get("row_key"),
            year=request.data.get("year"),
            month=request.data.get("month"),
        ).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class BudgetScenarioExportEndpoint(FinanceBaseView):
    model = BudgetScenario

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug, scenario_id):
        scenario = BudgetScenario.objects.get(id=scenario_id, workspace__slug=slug)
        export_format = request.query_params.get("format", "xlsx").lower()
        if export_format not in {"csv", "xlsx"}:
            return Response({"error": "format must be csv or xlsx"}, status=status.HTTP_400_BAD_REQUEST)

        forecast = _scenario_forecast_data(scenario)
        month_headers = [f'{item["year"]}-{item["month"]:02d}' for item in forecast["months"]]
        headers = ["Concept", "Type", "Entity", "Currency", *month_headers, "Total"]
        rows = [
            [
                line["label"],
                line["category"],
                line["entity_name"],
                line["currency"],
                *[month["amount"] for month in line["months"]],
                line["total"],
            ]
            for line in forecast["lines"]
        ]
        filename = f'{slugify(scenario.name) or "budget"}-{scenario.fiscal_year}.{export_format}'

        if export_format == "csv":
            output = StringIO()
            writer = csv.writer(output)
            writer.writerow(sanitize_csv_row(headers))
            for row in rows:
                writer.writerow(sanitize_csv_row(row))
            response = HttpResponse(output.getvalue(), content_type="text/csv; charset=utf-8")
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Budget"
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            sheet.freeze_panes = "E2"
            for cell in sheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="2F6B5F")
            for column in range(5, len(headers) + 1):
                for cell in sheet.iter_cols(min_col=column, max_col=column, min_row=2):
                    for item in cell:
                        item.number_format = '#,##0.00'
                        if item.value not in (None, ""):
                            item.value = Decimal(str(item.value))
            widths = {"A": 32, "B": 16, "C": 24, "D": 12}
            for column, width in widths.items():
                sheet.column_dimensions[column].width = width
            buffer = BytesIO()
            workbook.save(buffer)
            response = HttpResponse(
                buffer.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response


class ExpenseEndpoint(FinanceBaseView):
    serializer_class = ExpenseSerializer
    model = Expense

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        # Documents are serialized inline; prefetching keeps the list one query
        # instead of one per expense
        expenses = (
            Expense.objects.filter(workspace__slug=slug)
            .select_related("category")
            .prefetch_related("documents__asset")
        )

        category_ids = request.query_params.getlist("category")
        if "none" in category_ids:
            expenses = expenses.filter(category__isnull=True)
        elif category_ids:
            expenses = expenses.filter(category_id__in=category_ids)

        statuses = request.query_params.getlist("status")
        if statuses:
            expenses = expenses.filter(status__in=statuses)

        project_id = request.query_params.get("project")
        if project_id:
            expenses = expenses.filter(project_id=project_id)

        date_from = request.query_params.get("from")
        if date_from:
            expenses = expenses.filter(expense_date__gte=date_from)
        date_to = request.query_params.get("to")
        if date_to:
            expenses = expenses.filter(expense_date__lte=date_to)

        search = request.query_params.get("search")
        if search:
            expenses = expenses.filter(
                Q(vendor__icontains=search) | Q(description__icontains=search) | Q(reference__icontains=search)
            )
        return Response(ExpenseSerializer(expenses, many=True).data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        workspace = Workspace.objects.get(slug=slug)
        serializer = ExpenseSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save(workspace_id=workspace.id)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class ExpenseDetailEndpoint(FinanceBaseView):
    serializer_class = ExpenseSerializer
    model = Expense

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug, expense_id):
        expense = (
            Expense.objects.select_related("category")
            .prefetch_related("documents__asset")
            .get(id=expense_id, workspace__slug=slug)
        )
        return Response(ExpenseSerializer(expense).data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, expense_id):
        expense = Expense.objects.get(id=expense_id, workspace__slug=slug)
        serializer = ExpenseSerializer(expense, data=request.data, partial=True)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)
        serializer.save()
        return Response(serializer.data, status=status.HTTP_200_OK)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, expense_id):
        Expense.objects.get(id=expense_id, workspace__slug=slug).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ExpenseDocumentEndpoint(FinanceBaseView):
    """Attaches library files to an expense as supporting documents.

    The upload itself goes through the file library's presigned POST, so by the
    time we get here the bytes are already in the bucket and the asset exists.
    This only records the link — which is what lets one expense carry several
    invoices.
    """

    model = ExpenseDocument

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug, expense_id):
        expense = Expense.objects.get(id=expense_id, workspace__slug=slug)
        asset_ids = request.data.get("asset_ids") or []
        if not isinstance(asset_ids, list) or not asset_ids:
            return Response({"error": "asset_ids is required"}, status=status.HTTP_400_BAD_REQUEST)

        # Only the workspace's own uploaded assets — never a raw id from the
        # client pointing at someone else's file
        assets = FileAsset.objects.filter(
            id__in=asset_ids,
            workspace_id=expense.workspace_id,
            is_uploaded=True,
            is_deleted=False,
        )
        for asset in assets:
            ExpenseDocument.objects.get_or_create(
                expense=expense,
                asset=asset,
                defaults={"workspace_id": expense.workspace_id},
            )

        expense.refresh_from_db()
        return Response(ExpenseSerializer(expense).data, status=status.HTTP_201_CREATED)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, expense_id, asset_id):
        # Detaches the document from the expense; the file itself stays in the
        # library, since it may be linked elsewhere or wanted on its own
        ExpenseDocument.objects.filter(
            expense_id=expense_id, asset_id=asset_id, workspace__slug=slug
        ).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class ExpenseDocumentViewEndpoint(FinanceBaseView):
    """Presigned URL for one of an expense's documents, for the viewer.

    Payments resolves its own URLs instead of borrowing the file library's
    download route: that route is gated by the library's feature flag, so a
    workspace running payments without it would find every invoice unviewable.
    Going through the link table also means only a document actually attached to
    an expense in this workspace resolves.
    """

    model = ExpenseDocument

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug, expense_id, asset_id):
        document = (
            ExpenseDocument.objects.select_related("asset")
            .filter(expense_id=expense_id, asset_id=asset_id, workspace__slug=slug)
            .first()
        )
        if document is None or document.asset_id is None:
            return Response({"error": "Document not found"}, status=status.HTTP_404_NOT_FOUND)

        storage = S3Storage(request=request)
        # inline: the viewer renders it in place; ?download=1 forces attachment
        disposition = "attachment" if request.query_params.get("download") else "inline"
        url = storage.generate_presigned_url(
            object_name=document.asset.asset.name,
            disposition=disposition,
            filename=(document.asset.attributes or {}).get("name"),
        )
        return Response({"url": url}, status=status.HTTP_200_OK)


class BudgetSummaryEndpoint(FinanceBaseView):
    """Budgeted vs spent per category for a period.

    Both sides are grouped by currency and never added across currencies — a
    total of "10,000" that silently mixes MXN and USD is a wrong number, not a
    rounding detail. Cancelled expenses are excluded; pending ones are reported
    separately so a category can show what is already committed but unpaid.
    """

    model = Budget

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        date_from = request.query_params.get("from")
        date_to = request.query_params.get("to")
        if not date_from or not date_to:
            return Response(
                {"error": "from and to are required (YYYY-MM-DD)"}, status=status.HTTP_400_BAD_REQUEST
            )

        budgets = Budget.objects.filter(
            workspace__slug=slug, period_start__lte=date_to, period_end__gte=date_from
        ).select_related("category")

        expenses = Expense.objects.filter(
            workspace__slug=slug, expense_date__gte=date_from, expense_date__lte=date_to
        ).exclude(status=Expense.Status.CANCELLED)

        # (category_id, currency) -> amounts
        rows = {}

        def row_for(category_id, category_name, currency):
            key = (str(category_id) if category_id else None, currency)
            if key not in rows:
                rows[key] = {
                    "category_id": str(category_id) if category_id else None,
                    "category_name": category_name,
                    "currency": currency,
                    "budgeted": 0,
                    "spent": 0,
                    "pending": 0,
                }
            return rows[key]

        for budget in budgets.values("category_id", "category__name", "currency").annotate(total=Sum("amount")):
            entry = row_for(budget["category_id"], budget["category__name"], budget["currency"])
            entry["budgeted"] = budget["total"]

        paid_or_pending = expenses.values("category_id", "category__name", "currency", "status").annotate(
            total=Sum("amount")
        )
        for expense in paid_or_pending:
            entry = row_for(expense["category_id"], expense["category__name"], expense["currency"])
            if expense["status"] == Expense.Status.PAID:
                entry["spent"] += expense["total"]
            else:
                entry["pending"] += expense["total"]

        results = []
        for entry in rows.values():
            entry["remaining"] = entry["budgeted"] - entry["spent"]
            # Emit money as strings. Handing a raw Decimal to the JSON renderer
            # turns it into a float, which is both lossy and inconsistent with
            # the serializers (DRF renders DecimalField as a string).
            for field in ("budgeted", "spent", "pending", "remaining"):
                entry[field] = str(Decimal(entry[field]).quantize(CENTS))
            results.append(entry)

        results.sort(key=lambda item: (item["category_name"] or "", item["currency"]))
        return Response(
            {"from": date_from, "to": date_to, "results": results},
            status=status.HTTP_200_OK,
        )
