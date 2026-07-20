import csv
import json
import re
import unicodedata
from datetime import date, datetime, time, timedelta
from decimal import Decimal, InvalidOperation
from io import BytesIO, StringIO

from django.db import connection, transaction
from django.db.models import Count, Exists, OuterRef, Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.dateparse import parse_date
from openpyxl import Workbook, load_workbook
from rest_framework import status
from rest_framework.negotiation import DefaultContentNegotiation
from rest_framework.response import Response

from plane.app.permissions import ROLE, allow_permission
from plane.app.serializers import (
    MusicCompanySerializer,
    MusicGenreSerializer,
    MusicPartySerializer,
    MusicReleaseSerializer,
    MusicTrackSerializer,
)
from plane.db.models import (
    FileAsset,
    MusicCompany,
    MusicCredit,
    MusicDistribution,
    MusicGenre,
    MusicImportRun,
    MusicLink,
    MusicParty,
    MusicRelease,
    MusicReleaseArtist,
    MusicReleaseTrack,
    MusicTrack,
    MusicTrackGenre,
    Workspace,
    WorkspaceFeature,
)
from plane.utils.workspace_feature import is_workspace_feature_enabled

from ..base import BaseAPIView


class MusicBaseView(BaseAPIView):
    def initial(self, request, *args, **kwargs):
        super().initial(request, *args, **kwargs)
        slug = kwargs.get("slug")
        if slug and not is_workspace_feature_enabled(WorkspaceFeature.FeatureKey.MUSIC_CATALOG, slug=slug):
            self.permission_denied(request, message="Music catalog is not enabled for this workspace")


class MusicReportContentNegotiation(DefaultContentNegotiation):
    """Keep DRF's reserved format parameter from rejecting file exports."""

    def filter_renderers(self, renderers, format):
        if format in ("csv", "xlsx"):
            return renderers
        return super().filter_renderers(renderers, format)


ARTIST_SEPARATOR = re.compile(
    r"\s*(?:;|\||,|&|\band\b|\by\b|\bfeat(?:uring)?\.?\b|\bf\.?t\.?[,]?\b)\s*",
    re.IGNORECASE,
)

MONTHS = {
    "enero": "january", "febrero": "february", "marzo": "march", "abril": "april",
    "mayo": "may", "junio": "june", "julio": "july", "agosto": "august",
    "septiembre": "september", "setiembre": "september", "octubre": "october",
    "noviembre": "november", "diciembre": "december",
}


def _plain(value):
    return "".join(
        char for char in unicodedata.normalize("NFKD", str(value or "").lower()) if not unicodedata.combining(char)
    )


def _split(value):
    return [part.strip(" .") for part in ARTIST_SEPARATOR.split(str(value or "")) if part.strip(" .")]


def _prune(queryset):
    """Soft-delete relations one by one (keeps Plane's soft-delete semantics;
    `_restore_or_create` can revive them on a later import)."""
    for instance in queryset:
        instance.delete()


def _choice(raw, choices, default):
    """Normalize a raw cell to a TextChoices VALUE. Matches the raw string
    against choice values and labels case-insensitively so a CSV that says
    "Released" is stored as the canonical "RELEASED" (filters compare values)."""
    text = str(raw or "").strip()
    if not text:
        return default
    lowered = text.lower()
    for value, label in choices:
        if lowered in (str(value).lower(), str(label).lower()):
            return value
    return text


def _restore_or_create(model, workspace, lookup, values=None):
    """Revive a soft-deleted unique relation instead of inserting a duplicate."""
    values = values or {}
    instance = model.all_objects.filter(**lookup).order_by("-created_at").first()
    if instance:
        instance.deleted_at = None
        instance.workspace = workspace
        for field, value in values.items():
            setattr(instance, field, value)
        instance.save()
        return instance
    return model.objects.create(workspace=workspace, **lookup, **values)


def _date(value, year_only=False):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        value = value.date()
    if isinstance(value, date):
        if year_only and value.year < 1910:
            serial = (value - date(1899, 12, 30)).days
            if 1900 <= serial <= 2100:
                return date(serial, 1, 1)
        return value
    text = str(value).strip()
    if re.fullmatch(r"\d{4}(?:\.0)?", text):
        year = int(float(text))
        if 1900 <= year <= 2100:
            return date(year, 1, 1)
    parsed = parse_date(text)
    if parsed:
        return _date(parsed, year_only=year_only)
    normalized = _plain(text).replace(" de ", " ")
    for spanish, english in MONTHS.items():
        normalized = re.sub(rf"\b{spanish}\b", english, normalized)
    formats = (
        "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%m-%d-%Y", "%Y/%m/%d",
        "%d.%m.%Y", "%d %B %Y", "%B %d %Y", "%B %d, %Y", "%B %Y",
    )
    for fmt in formats:
        try:
            return _date(datetime.strptime(normalized, fmt).date(), year_only=year_only)
        except ValueError:
            continue
    return None


def _mapped_date(row, mapping, key):
    columns = _columns_for(mapping, key)
    column = columns[0] if columns else ""
    return _date(row.get(column), year_only=any(token in _plain(column) for token in ("year", "ano")))


def _duration_ms(value):
    if value in (None, ""):
        return None

    if isinstance(value, timedelta):
        return max(0, round(value.total_seconds() * 1000))
    if isinstance(value, (datetime, time)):
        return round(
            (value.hour * 3600 + value.minute * 60 + value.second + value.microsecond / 1_000_000) * 1000
        )

    text = str(value).strip()
    clock = re.search(r"(?<!\d)(\d{1,3}):(\d{1,2})(?::(\d{1,2}(?:\.\d+)?))?", text)
    if clock:
        first, second, third = clock.groups()
        seconds = (
            int(first) * 3600 + int(second) * 60 + float(third)
            if third is not None
            else int(first) * 60 + float(second)
        )
        return max(0, round(seconds * 1000))

    normalized = _plain(text).replace(",", ".")
    unit_parts = re.findall(
        r"(\d+(?:\.\d+)?)\s*(horas?|hrs?|h|minutos?|mins?|min|m|segundos?|segs?|secs?|seg|sec|s)\b",
        normalized,
    )
    if unit_parts:
        hours = sum(float(number) for number, unit in unit_parts if unit.startswith(("h", "hora")))
        minutes = sum(float(number) for number, unit in unit_parts if unit.startswith(("m", "minuto")))
        seconds = sum(float(number) for number, unit in unit_parts if unit.startswith(("s", "seg")))
        return max(0, round((hours * 3600 + minutes * 60 + seconds) * 1000))

    try:
        number = float(normalized)
        # Excel stores formatted times as a fraction of one day.
        if 0 < number < 1:
            return round(number * 86_400_000)
        return round(number * 1000) if number < 36000 else round(number)
    except ValueError:
        return None


def _decimal(value):
    try:
        return Decimal(str(value).replace("%", "").strip()) if value not in (None, "") else None
    except InvalidOperation:
        return None


def _bool(value):
    return str(value).strip().lower() in {"1", "true", "yes", "si", "sí", "y"}


def _columns_for(mapping, key):
    """Columns mapped to a field. Multi-capable fields (links, credits…) may
    map SEVERAL columns: the mapping value is then a list instead of a str."""
    value = mapping.get(key)
    if not value:
        return []
    return [column for column in (value if isinstance(value, list) else [value]) if column]


def _mapped(row, mapping, key, default=""):
    columns = _columns_for(mapping, key)
    column = columns[0] if columns else None
    return str(row.get(column, default)).strip() if column and row.get(column) is not None else default


def _mapped_many(row, mapping, key):
    """Non-empty values across EVERY column mapped to the field."""
    values = []
    for column in _columns_for(mapping, key):
        raw = row.get(column)
        if raw is None:
            continue
        text = str(raw).strip()
        if text:
            values.append(text)
    return values


def _import_error_message(exc):
    message = str(exc)
    if "music_tracks.parent_track_id" in message and "does not exist" in message:
        return (
            "The music catalog database is not ready. Apply migration "
            "db.0136_musictrack_parent_track and restart the API before importing."
        )
    return message


def _jsonable_import_value(value):
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _import_error_detail(index, exc, row, mapping):
    message = _import_error_message(exc)
    field = None
    code = "ROW_VALIDATION_ERROR"
    if message == "Track title is empty":
        field = "track.title"
        code = "REQUIRED_FIELD"
    elif message.startswith("Duplicate track:"):
        field = "track.isrc" if mapping.get("track.isrc") else "track.title"
        code = "DUPLICATE"
    elif message.startswith("The music catalog database is not ready"):
        code = "DATABASE_NOT_READY"
    field_columns = _columns_for(mapping, field) if field else []
    column = field_columns[0] if field_columns else None
    return {
        "row": index,
        "code": code,
        "field": field,
        "column": column,
        "value": _jsonable_import_value(row.get(column)) if column else None,
        "message": message,
        "row_data": {
            str(key): _jsonable_import_value(value) for key, value in list(row.items())[:12]
        },
    }


def _apply_row_overrides(row, mapping, overrides):
    if not overrides:
        return row, mapping
    next_row = dict(row)
    next_mapping = dict(mapping)
    for field, value in overrides.items():
        override_column = f"__override__{field}"
        next_row[override_column] = value
        next_mapping[field] = override_column
    return next_row, next_mapping


def _music_schema_ready():
    try:
        with connection.cursor() as cursor:
            columns = {
                column.name for column in connection.introspection.get_table_description(cursor, "music_tracks")
            }
        return "parent_track_id" in columns
    except Exception:
        return False


def _unique_headers(values):
    headers, counts = [], {}
    for index, value in enumerate(values, start=1):
        base = str(value or "").strip() or f"Column {index}"
        counts[base] = counts.get(base, 0) + 1
        headers.append(base if counts[base] == 1 else f"{base} ({counts[base]})")
    return headers


def _header_index(rows):
    best_index, best_score = 0, -1
    for index, row in enumerate(rows[:25]):
        values = [str(value).strip() for value in row if value not in (None, "")]
        words = sum(bool(re.search(r"[A-Za-z\u00c0-\u024f]", value)) for value in values)
        score = words * 3 + len(set(values)) - index * 0.05
        if len(values) >= 2 and score > best_score:
            best_index, best_score = index, score
    return best_index


def _read_table(upload, sheet_name=None):
    content = upload.read()
    filename = upload.name.lower()
    sheets = []
    if filename.endswith((".xlsx", ".xlsm")):
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
        sheets = workbook.sheetnames
        worksheet = workbook[sheet_name] if sheet_name in sheets else workbook[workbook.sheetnames[0]]
        raw_rows = [list(row) for row in worksheet.iter_rows(values_only=True)]
    else:
        decoded = None
        for encoding in ("utf-8-sig", "cp1252", "latin-1"):
            try:
                decoded = content.decode(encoding)
                break
            except UnicodeDecodeError:
                continue
        sample = (decoded or "")[:8192]
        try:
            delimiter = csv.Sniffer().sniff(sample, delimiters=",;\t|").delimiter
        except csv.Error:
            delimiter = ","
        raw_rows = list(csv.reader(StringIO(decoded or ""), delimiter=delimiter))
    if not raw_rows or not any(any(value not in (None, "") for value in row) for row in raw_rows):
        return [], [], sheets, 0
    header_row = _header_index(raw_rows)
    width = max(len(row) for row in raw_rows)
    headers = _unique_headers(list(raw_rows[header_row]) + [None] * (width - len(raw_rows[header_row])))
    rows = [
        {headers[index]: value for index, value in enumerate(list(row) + [None] * (width - len(row)))}
        for row in raw_rows[header_row + 1 :]
        if any(value not in (None, "") for value in row)
    ]
    return headers, rows, sheets, header_row + 1


class _StoredUpload:
    """File-like (.read() + .name) over a stored import asset so every import
    endpoint can consume a saved file exactly like a fresh browser upload."""

    def __init__(self, content, name):
        self._content = content
        self.name = name

    def read(self):
        return self._content

    def seek(self, _position):
        return None


def _asset_upload(slug, asset_id):
    """Loads a saved MUSIC_CATALOG import asset as an upload-like object (for
    re-imports the file is never re-uploaded by the browser)."""
    from plane.settings.storage import S3Storage

    asset = FileAsset.objects.get(
        id=asset_id,
        workspace__slug=slug,
        entity_type=FileAsset.EntityTypeContext.MUSIC_CATALOG,
        is_deleted=False,
    )
    storage = S3Storage()
    obj = storage.s3_client.get_object(Bucket=storage.aws_storage_bucket_name, Key=asset.asset.name)
    name = (asset.attributes or {}).get("name") or asset.asset.name
    return asset, _StoredUpload(obj["Body"].read(), name)


IMPORT_ALIASES = {
    "track.title": ("track", "song", "song name", "titulo", "cancion"),
    "track.isrc": ("isrc", "isrc audio", "isrc song"),
    "track.isrc_video": ("isrc video", "video isrc"),
    "track.catalog": ("catalog", "catalogo", "catalog number", "no cat", "cat"),
    "track.upc": ("upc", "ean", "barcode"),
    "track.release_date": ("track release date", "fecha lanzamiento cancion"),
    "track.duration_ms": ("duration", "duracion", "duracion tipo", "time"),
    "release.title": ("album", "album name", "titulo album single lp ep"),
    "release.upc": ("upc",),
    "release.catalog_number": ("catalog", "catalog number", "product catalog"),
    "release.release_date": ("album release date", "release date", "fecha ano", "fecha"),
    "artists": ("artist", "artists", "artista", "main artist"),
    "writers": ("writer", "writers", "escritor", "escritores", "songwriter", "songwriters"),
    "record_label": ("label", "label name", "licencia", "statement"),
    "track.video_url": ("video url", "video link", "link video", "url video", "youtube", "music video url"),
    "track.video_release_date": ("video release date", "fecha video", "fecha lanzamiento video"),
    "track.streaming_url": ("streaming url", "spotify", "spotify url", "enlace", "song url"),
}

IMPORT_FIELDS = [
    "track.title", "track.version", "track.isrc", "track.isrc_video", "track.catalog", "track.upc", "track.kind",
    "track.status", "track.release_date", "track.original_release_date", "track.duration_ms",
    "track.country_of_recording", "track.language", "track.explicit", "track.ownership",
    "track.us_publishing_obligations", "track.recoupment", "track.digital_format", "track.lyrics",
    "track.aggregator_percentage", "track.distributor_percentage", "track.record_label_percentage",
    "track.artist_percentage", "track.writer_percentage", "release.title", "release.type", "release.upc",
    "release.catalog_number", "release.release_date", "release.p_line", "release.c_line", "artists",
    "featured_artists", "writers", "authors", "composers", "producers", "recording_engineers", "mixers",
    "mastering_engineers", "legal_representatives", "genres", "record_label", "aggregator", "distributor",
    # Content-detected columns: a column of URLs maps here even when its
    # header says nothing useful (the assistant decides by LOOKING at values)
    "track.video_url", "track.video_release_date", "track.streaming_url",
]


def _link_platform(url):
    """Readable platform name from a URL's host ("YouTube", "Spotify"…)."""
    match = re.search(r"https?://(?:www\.)?([^/:]+)", str(url or ""))
    if not match:
        return ""
    host = match.group(1)
    known = {
        "youtube.com": "YouTube", "youtu.be": "YouTube", "open.spotify.com": "Spotify",
        "music.apple.com": "Apple Music", "deezer.com": "Deezer", "tidal.com": "Tidal",
        "soundcloud.com": "SoundCloud", "music.amazon.com": "Amazon Music",
    }
    for domain, name in known.items():
        if host == domain or host.endswith("." + domain):
            return name
    return host.split(".")[0].capitalize()


# Typed canonical fields: raw values that fail these parsers are "variables"
# ("ringtone" in a duration column) the user must resolve via value_overrides
UNPARSED_CHECKS = {
    "track.duration_ms": _duration_ms,
    "track.release_date": _date,
    "track.original_release_date": _date,
    "track.video_release_date": _date,
    "release.release_date": _date,
    "track.aggregator_percentage": _decimal,
    "track.distributor_percentage": _decimal,
    "track.record_label_percentage": _decimal,
    "track.artist_percentage": _decimal,
    "track.writer_percentage": _decimal,
}

# value_overrides sentinel: rows whose raw value maps to this are not imported
SKIP_ROW = "__SKIP_ROW__"

# Fields that accept SEVERAL source columns (mapping value = list of columns):
# link columns and people/genre/company columns can appear more than once.
MULTI_COLUMN_FIELDS = [
    "artists", "featured_artists", "writers", "authors", "composers", "producers", "recording_engineers",
    "mixers", "mastering_engineers", "legal_representatives", "genres", "record_label",
    "aggregator", "distributor", "track.video_url", "track.streaming_url",
]

# Example of the expected format per typed field, shown when the user decides
# to replace an unparseable "variable" token
FIELD_FORMAT_EXAMPLES = {
    "track.duration_ms": "3:16",
    "track.release_date": "2023-05-01",
    "track.original_release_date": "2023-05-01",
    "track.video_release_date": "2023-05-01",
    "release.release_date": "2023-05-01",
    "track.aggregator_percentage": "50",
    "track.distributor_percentage": "50",
    "track.record_label_percentage": "50",
    "track.artist_percentage": "50",
    "track.writer_percentage": "50",
}


def _column_samples(headers, rows, limit=10):
    """Per column: fill count and up to `limit` non-empty examples scanned over
    the WHOLE file — a column empty for thousands of rows and populated later
    must still be classifiable (by the AI mapper or the user)."""
    samples = {}
    for header in headers:
        examples = []
        non_empty = 0
        for row in rows:
            value = row.get(header)
            if value in (None, ""):
                continue
            non_empty += 1
            if len(examples) < limit:
                text = value.isoformat() if isinstance(value, (date, datetime)) else str(value)
                examples.append(text[:120])
        samples[header] = {"non_empty": non_empty, "total": len(rows), "examples": examples}
    return samples


def _collect_unparsed(row, mapping):
    """{canonical_field: raw_token} for typed columns whose value parses to
    nothing — the import would silently drop them without user input."""
    found = {}
    for field, parser in UNPARSED_CHECKS.items():
        for column in _columns_for(mapping, field):
            raw = row.get(column)
            if raw in (None, ""):
                continue
            if parser(raw) is None:
                found[field] = str(raw).strip()
                break
    return found


def _apply_overrides(row, mapping, value_overrides):
    """Returns (row_with_replacements, skip_row). Overrides are keyed by
    canonical field → {raw_token(lower): replacement}; empty replacement
    imports the row with that cell blank, SKIP_ROW drops the whole row."""
    if not value_overrides:
        return row, False
    patched = dict(row)
    for field, tokens in value_overrides.items():
        if not isinstance(tokens, dict):
            continue
        for column in _columns_for(mapping, field):
            raw = patched.get(column)
            if raw in (None, ""):
                continue
            replacement = tokens.get(str(raw).strip().lower())
            if replacement is None:
                continue
            if replacement == SKIP_ROW:
                return patched, True
            patched[column] = replacement
    return patched, False


OUTCOME_ACTION = {"created": "CREATED", "updated": "UPDATED", "skipped": "PRESERVED"}


def _record_import_run(workspace, *, file_asset, source_name, source, sheet, rules, summary, touched):
    """Persists the provenance of one applied import: the run itself plus a
    link per affected track (created/updated/preserved, with its row)."""
    from plane.db.models import MusicImportRun, MusicTrackImport

    run = MusicImportRun.objects.create(
        workspace=workspace,
        file_asset=file_asset,
        source_name=(source_name or "")[:500],
        source=source,
        sheet=sheet or "",
        rules=rules,
        summary=summary,
    )
    # unique (track, run): if several rows touched the same track keep the last
    by_track = {}
    for track_id, outcome, row_number in touched:
        by_track[track_id] = (OUTCOME_ACTION.get(outcome, "PRESERVED"), row_number)
    MusicTrackImport.objects.bulk_create(
        [
            MusicTrackImport(
                workspace=workspace,
                track_id=track_id,
                import_run=run,
                action=action,
                row_number=row_number,
            )
            for track_id, (action, row_number) in by_track.items()
        ],
        ignore_conflicts=True,
    )
    return run


def _infer_mapping(headers, fields):
    key = lambda value: re.sub(r"[^a-z0-9]+", " ", _plain(value)).strip()
    normalized = {key(header): header for header in headers}
    result = {}
    for field in fields:
        candidates = (field, field.split(".")[-1], *IMPORT_ALIASES.get(field, ()))
        for candidate in candidates:
            wanted = key(candidate)
            match = next((header for key, header in normalized.items() if key == wanted), None)
            if match:
                result[field] = match
                break
    return result


def _party(workspace, name, kind=MusicParty.Kind.ARTIST):
    party = MusicParty.objects.filter(workspace=workspace, kind=kind, display_name__iexact=name).first()
    return party or MusicParty.objects.create(workspace=workspace, kind=kind, display_name=name)


def _genre(workspace, name):
    genre = MusicGenre.objects.filter(workspace=workspace, name__iexact=name).first()
    return genre or MusicGenre.objects.create(workspace=workspace, name=name)


def _filter_releases(queryset, params):
    search = params.get("search")
    if search:
        queryset = queryset.filter(
            Q(title__icontains=search)
            | Q(upc__icontains=search)
            | Q(catalog_number__icontains=search)
            | Q(artist_links__party__display_name__icontains=search)
        )
    if params.get("type"):
        queryset = queryset.filter(release_type=params["type"])
    if params.get("status"):
        queryset = queryset.filter(status=params["status"])
    if params.get("artist"):
        queryset = queryset.filter(artist_links__party_id=params["artist"])
    if params.get("year"):
        queryset = queryset.filter(release_date__year=params["year"])
    if params.get("from"):
        queryset = queryset.filter(release_date__gte=params["from"])
    if params.get("to"):
        queryset = queryset.filter(release_date__lte=params["to"])
    if params.get("upc"):
        queryset = queryset.filter(upc__iexact=params["upc"])
    return queryset.distinct()


def _filter_tracks(queryset, params):
    search = params.get("search")
    if search:
        queryset = queryset.filter(
            Q(title__icontains=search)
            | Q(isrc__icontains=search)
            | Q(isrc_video__icontains=search)
            | Q(catalog__icontains=search)
            | Q(credits__party__display_name__icontains=search)
            | Q(videos__title__icontains=search)
            | Q(videos__isrc_video__icontains=search)
        )
    def _values(key):
        return [value for value in params.get(key, "").split(",") if value]

    for field in ("kind", "isrc", "isrc_video", "language", "country_of_recording"):
        if params.get(field):
            queryset = queryset.filter(**{field: params[field]})
    # Dynamic-entity filters accept comma-separated lists (multi-select UI)
    if params.get("status"):
        # iexact per value so legacy rows stored with the label ("Released")
        # still match the canonical enum sent by the UI ("RELEASED").
        status_q = Q()
        for value in _values("status"):
            status_q |= Q(status__iexact=value)
        queryset = queryset.filter(status_q)
    if params.get("artist"):
        queryset = queryset.filter(credits__party_id__in=_values("artist"))
    if params.get("genre"):
        queryset = queryset.filter(genre_links__genre_id__in=_values("genre"))
    if params.get("release"):
        queryset = queryset.filter(release_links__release_id__in=_values("release"))
    if params.get("year"):
        years = [value for value in _values("year") if value.isdigit()]
        if years:
            queryset = queryset.filter(
                Q(release_date__year__in=years) | Q(release_links__release__release_date__year__in=years)
            )
    if params.get("has_lyrics") == "true":
        queryset = queryset.exclude(lyrics="")
    if params.get("from"):
        queryset = queryset.filter(release_date__gte=params["from"])
    if params.get("to"):
        queryset = queryset.filter(release_date__lte=params["to"])
    if params.get("has_video") == "true":
        # A real music video carries at least one identifier or a video URL —
        # a bare child track (e.g. a stray import row) is not "having video".
        videos = MusicTrack.objects.filter(
            parent_track_id=OuterRef("pk"),
            kind=MusicTrack.Kind.MUSIC_VIDEO,
        ).filter(
            Q(links__kind=MusicLink.Kind.MUSIC_VIDEO)
            | ~Q(isrc_video="")
            | ~Q(upc="")
            | ~Q(catalog="")
        )
        queryset = queryset.filter(Exists(videos))
    if params.get("has_links") == "true":
        queryset = queryset.filter(Exists(MusicLink.objects.filter(track_id=OuterRef("pk"))))
    if params.get("video_from"):
        queryset = queryset.filter(videos__release_date__gte=params["video_from"])
    if params.get("video_to"):
        queryset = queryset.filter(videos__release_date__lte=params["video_to"])
    if params.get("company"):
        queryset = queryset.filter(distributions__company_id__in=_values("company"))
    if params.get("ids"):
        queryset = queryset.filter(id__in=_values("ids"))
    if params.get("import_run"):
        queryset = queryset.filter(import_links__import_run_id=params["import_run"])
    if params.get("import_file"):
        # Any run of those source files (re-imports of the same upload included)
        queryset = queryset.filter(import_links__import_run__file_asset_id__in=_values("import_file"))
    return queryset.distinct()


class MusicReleaseEndpoint(MusicBaseView):
    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        queryset = MusicRelease.objects.filter(workspace__slug=slug).prefetch_related("artist_links__party").annotate(
            track_count=Count("track_links", distinct=True)
        )
        return Response(MusicReleaseSerializer(_filter_releases(queryset, request.query_params), many=True).data)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        workspace = Workspace.objects.get(slug=slug)
        serializer = MusicReleaseSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        release = serializer.save(workspace=workspace)
        for artist_id in request.data.get("artist_ids", []):
            _restore_or_create(
                MusicReleaseArtist,
                workspace,
                {"release": release, "party_id": artist_id, "role": MusicReleaseArtist.Role.PRIMARY},
            )
        return Response(MusicReleaseSerializer(release).data, status=status.HTTP_201_CREATED)


class MusicReleaseDetailEndpoint(MusicBaseView):
    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug, release_id):
        release = MusicRelease.objects.prefetch_related("artist_links__party").get(id=release_id, workspace__slug=slug)
        return Response(MusicReleaseSerializer(release).data)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, release_id):
        release = MusicRelease.objects.get(id=release_id, workspace__slug=slug)
        serializer = MusicReleaseSerializer(release, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        if "artist_ids" in request.data:
            release.artist_links.all().delete()
            for artist_id in request.data["artist_ids"]:
                _restore_or_create(
                    MusicReleaseArtist,
                    release.workspace,
                    {"release": release, "party_id": artist_id, "role": MusicReleaseArtist.Role.PRIMARY},
                )
        return Response(MusicReleaseSerializer(release).data)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, release_id):
        MusicRelease.objects.get(id=release_id, workspace__slug=slug).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MusicTrackEndpoint(MusicBaseView):
    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        queryset = MusicTrack.objects.filter(workspace__slug=slug).prefetch_related(
            "credits__party", "genre_links__genre", "release_links__release", "links",
            "distributions__company", "videos__links",
        )
        if request.query_params.get("songs_only") == "true":
            queryset = queryset.filter(parent_track__isnull=True)
        queryset = _filter_tracks(queryset, request.query_params)

        # Lightweight id list so the UI can select EVERYTHING matching the
        # active filters without serializing thousands of full tracks.
        if request.query_params.get("ids_only") == "true":
            return Response({"ids": [str(pk) for pk in queryset.values_list("id", flat=True)]})

        # Keep the original array response for existing consumers that do not request a page.
        if "page" not in request.query_params:
            return Response(MusicTrackSerializer(queryset, many=True).data)

        try:
            requested_page = max(1, int(request.query_params.get("page", 1)))
            page_size = min(100, max(1, int(request.query_params.get("page_size", 25))))
        except (TypeError, ValueError):
            requested_page, page_size = 1, 25

        total = queryset.count()
        page_count = max(1, (total + page_size - 1) // page_size)
        page = min(requested_page, page_count)
        start = (page - 1) * page_size
        results = MusicTrackSerializer(queryset[start : start + page_size], many=True).data
        return Response(
            {
                "results": results,
                "total": total,
                "page": page,
                "page_size": page_size,
                "requested_page": requested_page,
            }
        )

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        workspace = Workspace.objects.get(slug=slug)
        serializer = MusicTrackSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        track = serializer.save(workspace=workspace)
        self._sync_relations(track, workspace, request.data)
        return Response(MusicTrackSerializer(track).data, status=status.HTTP_201_CREATED)

    @staticmethod
    def _sync_relations(track, workspace, data):
        if "releases" in data:
            track.release_links.all().delete()
            for release in data.get("releases", []):
                _restore_or_create(
                    MusicReleaseTrack,
                    workspace,
                    {"track": track, "release_id": release["id"]},
                    {
                        "disc_number": release.get("disc_number", 1),
                        "track_number": release.get("track_number", 1),
                    },
                )
        if "genre_ids" in data:
            track.genre_links.all().delete()
            for genre_id in data.get("genre_ids", []):
                _restore_or_create(
                    MusicTrackGenre,
                    workspace,
                    {"track": track, "genre_id": genre_id},
                )
        if "credit_entries" in data:
            track.credits.all().delete()
            for credit in data.get("credit_entries", []):
                _restore_or_create(
                    MusicCredit,
                    workspace,
                    {"track": track, "party_id": credit["party_id"], "role": credit["role"]},
                    {
                        "percentage": credit.get("percentage") or None,
                        "publishing_share": credit.get("publishing_share") or None,
                        "territory": credit.get("territory", ""),
                        "notes": credit.get("notes", ""),
                    },
                )
        if "link_entries" in data:
            track.links.all().delete()
            for link in data.get("link_entries", []):
                MusicLink.objects.create(
                    workspace=workspace,
                    track=track,
                    kind=link.get("kind", MusicLink.Kind.OTHER),
                    platform=link.get("platform", ""),
                    name=link.get("name") or link.get("platform") or "Link",
                    url=link["url"],
                    isrc=link.get("isrc", ""),
                )
        if "distribution_entries" in data:
            track.distributions.all().delete()
            for entry in data.get("distribution_entries", []):
                MusicDistribution.objects.create(
                    workspace=workspace,
                    track=track,
                    company_id=entry["company_id"],
                    percentage=entry.get("percentage") or None,
                    territory=entry.get("territory", ""),
                    valid_from=entry.get("valid_from") or None,
                    valid_to=entry.get("valid_to") or None,
                )
        if "video_entries" in data:
            retained = []
            for entry in data.get("video_entries", []):
                video_id = entry.get("id")
                video = track.videos.filter(id=video_id).first() if video_id else None
                values = {
                    "title": entry.get("title") or track.title,
                    "version": entry.get("version", ""),
                    "status": entry.get("status", MusicTrack.Status.DRAFT),
                    "kind": MusicTrack.Kind.MUSIC_VIDEO,
                    "isrc_video": str(entry.get("isrc", "")).upper().replace("-", ""),
                    "upc": entry.get("upc", ""),
                    "catalog": entry.get("catalog", ""),
                    "release_date": entry.get("release_date") or None,
                    "duration_ms": entry.get("duration_ms") or None,
                    "cover_url": entry.get("cover_url", ""),
                }
                if video:
                    for key, value in values.items():
                        setattr(video, key, value)
                    video.save()
                else:
                    video = MusicTrack.objects.create(workspace=workspace, parent_track=track, **values)
                retained.append(video.id)
                video.links.filter(kind=MusicLink.Kind.MUSIC_VIDEO).delete()
                if entry.get("video_url"):
                    MusicLink.objects.create(
                        workspace=workspace,
                        track=video,
                        kind=MusicLink.Kind.MUSIC_VIDEO,
                        platform=entry.get("platform", ""),
                        name=entry.get("name") or "Music video",
                        url=entry["video_url"],
                        isrc=values["isrc_video"],
                    )
            track.videos.exclude(id__in=retained).delete()


class MusicTrackBulkDeleteEndpoint(MusicBaseView):
    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        track_ids = request.data.get("track_ids", [])
        if not isinstance(track_ids, list) or not track_ids:
            return Response({"error": "track_ids must contain at least one song"}, status=status.HTTP_400_BAD_REQUEST)

        with transaction.atomic():
            tracks = list(MusicTrack.objects.filter(workspace__slug=slug, id__in=track_ids))
            for track in tracks:
                track.delete()

        return Response({"deleted": len(tracks), "not_found": len(track_ids) - len(tracks)})


class MusicTrackDetailEndpoint(MusicBaseView):
    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug, track_id):
        track = MusicTrack.objects.prefetch_related(
            "credits__party", "genre_links__genre", "release_links__release", "links",
            "distributions__company", "videos__links",
        ).get(id=track_id, workspace__slug=slug)
        return Response(MusicTrackSerializer(track).data)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, track_id):
        track = MusicTrack.objects.get(id=track_id, workspace__slug=slug)
        serializer = MusicTrackSerializer(track, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        MusicTrackEndpoint._sync_relations(track, track.workspace, request.data)
        return Response(MusicTrackSerializer(track).data)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, track_id):
        MusicTrack.objects.get(id=track_id, workspace__slug=slug).delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class _NamedEndpoint(MusicBaseView):
    model = None
    serializer = None

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        return Response(self.serializer(self.model.objects.filter(workspace__slug=slug), many=True).data)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        serializer = self.serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save(workspace=Workspace.objects.get(slug=slug))
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class _NamedDetailEndpoint(MusicBaseView):
    model = None
    serializer = None
    lookup_kwarg = "resource_id"

    def _instance(self, slug, **kwargs):
        return self.model.objects.get(id=kwargs[self.lookup_kwarg], workspace__slug=slug)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def patch(self, request, slug, **kwargs):
        instance = self._instance(slug, **kwargs)
        serializer = self.serializer(instance, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def delete(self, request, slug, **kwargs):
        instance = self._instance(slug, **kwargs)
        if isinstance(instance, MusicParty) and (instance.track_credits.exists() or instance.release_links.exists()):
            return Response(
                {"error": "This person or artist is used by tracks or releases. Disable it instead of deleting it."},
                status=status.HTTP_409_CONFLICT,
            )
        instance.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class MusicPartyEndpoint(_NamedEndpoint):
    model = MusicParty
    serializer = MusicPartySerializer


class MusicGenreEndpoint(_NamedEndpoint):
    model = MusicGenre
    serializer = MusicGenreSerializer


class MusicCompanyEndpoint(_NamedEndpoint):
    model = MusicCompany
    serializer = MusicCompanySerializer


class MusicPartyDetailEndpoint(_NamedDetailEndpoint):
    model = MusicParty
    serializer = MusicPartySerializer
    lookup_kwarg = "party_id"


class MusicGenreDetailEndpoint(_NamedDetailEndpoint):
    model = MusicGenre
    serializer = MusicGenreSerializer
    lookup_kwarg = "genre_id"


class MusicCompanyDetailEndpoint(_NamedDetailEndpoint):
    model = MusicCompany
    serializer = MusicCompanySerializer
    lookup_kwarg = "company_id"


class MusicCatalogOptionsEndpoint(MusicBaseView):
    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        return Response(
            {
                "release_types": MusicRelease.Type.choices,
                "release_statuses": MusicRelease.Status.choices,
                "track_kinds": MusicTrack.Kind.choices,
                "track_statuses": MusicTrack.Status.choices,
                "credit_roles": MusicCredit.Role.choices,
                "party_kinds": MusicParty.Kind.choices,
                "company_kinds": MusicCompany.Kind.choices,
                "import_fields": IMPORT_FIELDS,
                "multi_fields": MULTI_COLUMN_FIELDS,
                # One entry per FILE (several runs of the same upload collapse
                # into it); files deleted from the library disappear here too.
                "import_files": self._import_files(slug),
            }
        )

    @staticmethod
    def _import_files(slug):
        files = {}
        runs = MusicImportRun.objects.filter(
            workspace__slug=slug, file_asset__isnull=False, file_asset__is_deleted=False
        ).order_by("-created_at")[:500]
        for run in runs:
            asset_id = str(run.file_asset_id)
            entry = files.get(asset_id)
            if entry is None:
                files[asset_id] = {
                    "asset_id": asset_id,
                    "name": run.source_name,
                    "source": run.source,
                    "runs": 1,
                    "last_imported_at": run.created_at.isoformat(),
                }
            else:
                entry["runs"] += 1
        return list(files.values())[:100]


class MusicImportAssetEndpoint(MusicBaseView):
    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        assets = FileAsset.objects.filter(
            workspace__slug=slug,
            entity_type=FileAsset.EntityTypeContext.MUSIC_CATALOG,
            attributes__music_asset_kind="IMPORT_SOURCE",
            is_uploaded=True,
            is_deleted=False,
        ).order_by("-created_at")
        search = request.query_params.get("search", "").strip()
        if search:
            assets = assets.filter(attributes__name__icontains=search)
        assets = list(assets[:500])

        # Latest run per asset: its rules restore the full panel configuration
        # when the user chooses to re-import a saved file.
        last_runs = {}
        for run in MusicImportRun.objects.filter(
            file_asset_id__in=[asset.id for asset in assets]
        ).order_by("-created_at"):
            last_runs.setdefault(run.file_asset_id, run)

        def last_run_payload(asset):
            run = last_runs.get(asset.id)
            if run is None:
                return None
            return {
                "sheet": run.sheet or None,
                "rules": run.rules or {},
                "summary": run.summary or {},
                "imported_at": run.created_at.isoformat(),
            }

        return Response(
            {
                "results": [
                    {
                        "id": str(asset.id),
                        "name": (asset.attributes or {}).get("name") or "Import file",
                        "content_type": (asset.attributes or {}).get("type") or "application/octet-stream",
                        "size": asset.size,
                        "upload_source": (asset.attributes or {}).get("upload_source") or "manual",
                        "created_at": asset.created_at.isoformat(),
                        "last_run": last_run_payload(asset),
                    }
                    for asset in assets
                ]
            }
        )

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        if request.data.get("action") != "delete":
            return Response({"action": ["Unsupported action"]}, status=status.HTTP_400_BAD_REQUEST)
        asset_ids = request.data.get("asset_ids") or []
        if not isinstance(asset_ids, list) or not asset_ids:
            return Response({"asset_ids": ["Select at least one file"]}, status=status.HTTP_400_BAD_REQUEST)
        assets = FileAsset.objects.filter(
            id__in=asset_ids,
            workspace__slug=slug,
            entity_type=FileAsset.EntityTypeContext.MUSIC_CATALOG,
            attributes__music_asset_kind="IMPORT_SOURCE",
            is_deleted=False,
        )
        deleted = assets.count()
        assets.update(is_deleted=True, deleted_at=timezone.now())
        return Response({"deleted": deleted, "not_found": len(set(asset_ids)) - deleted})


class MusicImportPreviewEndpoint(MusicBaseView):
    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        upload = request.FILES.get("file")
        if not upload and request.data.get("asset_id"):
            try:
                _, upload = _asset_upload(slug, request.data["asset_id"])
            except FileAsset.DoesNotExist:
                return Response({"error": "Archivo guardado no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        if not upload:
            return Response({"file": ["Choose a CSV or XLSX file"]}, status=status.HTTP_400_BAD_REQUEST)
        try:
            headers, rows, sheets, header_row = _read_table(upload, request.data.get("sheet"))
        except Exception as exc:
            return Response({"error": f"Could not read spreadsheet: {exc}"}, status=status.HTTP_400_BAD_REQUEST)
        sample = [
            {key: value.isoformat() if isinstance(value, (date, datetime)) else value for key, value in row.items()}
            for row in rows[:8]
        ]
        inferred = _infer_mapping(headers, IMPORT_FIELDS)
        artist_column = inferred.get("artists")
        database_ready = _music_schema_ready()
        return Response(
            {
                "headers": headers,
                "rows": sample,
                "sheets": sheets,
                "selected_sheet": request.data.get("sheet") or (sheets[0] if sheets else None),
                "header_row": header_row,
                "total_rows": len(rows),
                "mapping": inferred,
                "column_samples": _column_samples(headers, rows),
                "multi_fields": MULTI_COLUMN_FIELDS,
                "format_examples": FIELD_FORMAT_EXAMPLES,
                "artist_examples": [
                    {"source": str(row.get(artist_column, "")), "detected": _split(row.get(artist_column, ""))}
                    for row in rows[:5]
                    if artist_column and row.get(artist_column)
                ],
                "database_ready": database_ready,
                "database_error": None if database_ready else "Apply migration db.0136_musictrack_parent_track before importing.",
            }
        )


class MusicImportAIMapEndpoint(MusicBaseView):
    """Optional AI mapping for the manual import panel: column samples from
    the whole file go to the Worker's model, which returns canonical_field →
    column(s). Deterministic validation happens here; the AI only suggests."""

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        from plane.utils.worker_client import WorkerTriggerError, ai_map_music_columns

        upload = request.FILES.get("file")
        if not upload and request.data.get("asset_id"):
            try:
                _, upload = _asset_upload(slug, request.data["asset_id"])
            except FileAsset.DoesNotExist:
                return Response({"error": "Archivo guardado no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        if not upload:
            return Response({"file": ["Choose a CSV or XLSX file"]}, status=status.HTTP_400_BAD_REQUEST)
        try:
            headers, rows, _sheets, _header_row = _read_table(upload, request.data.get("sheet"))
        except Exception as exc:
            return Response({"error": f"Could not read spreadsheet: {exc}"}, status=status.HTTP_400_BAD_REQUEST)
        try:
            data = ai_map_music_columns(_column_samples(headers, rows), IMPORT_FIELDS, MULTI_COLUMN_FIELDS)
        except WorkerTriggerError as exc:
            return Response({"error": str(exc)[:300]}, status=status.HTTP_502_BAD_GATEWAY)

        # Only real fields mapped to real columns survive
        header_set = set(headers)
        mapping = {}
        for field, value in (data.get("mapping") or {}).items():
            if field not in IMPORT_FIELDS:
                continue
            columns = [column for column in (value if isinstance(value, list) else [value]) if column in header_set]
            if not columns:
                continue
            mapping[field] = columns if field in MULTI_COLUMN_FIELDS and len(columns) > 1 else columns[0]
        return Response({"mapping": mapping, "model": data.get("model")}, status=status.HTTP_200_OK)


class MusicImportEndpoint(MusicBaseView):
    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def post(self, request, slug):
        upload = request.FILES.get("file")
        source_asset = None
        if not upload and request.data.get("asset_id"):
            # Re-import of a stored file: the browser never re-uploads it
            try:
                source_asset, upload = _asset_upload(slug, request.data["asset_id"])
            except FileAsset.DoesNotExist:
                return Response({"error": "Archivo guardado no encontrado"}, status=status.HTTP_404_NOT_FOUND)
        if not upload:
            return Response({"file": ["A CSV or XLSX file is required"]}, status=status.HTTP_400_BAD_REQUEST)
        try:
            mapping = json.loads(request.data.get("mapping", "{}"))
            defaults = json.loads(request.data.get("defaults", "{}"))
            _, rows, _, header_row = _read_table(upload, request.data.get("sheet"))
        except (json.JSONDecodeError, csv.Error, ValueError) as exc:
            return Response({"error": f"Invalid spreadsheet: {exc}"}, status=status.HTTP_400_BAD_REQUEST)
        if not mapping.get("track.title"):
            return Response({"mapping": ["track.title must be mapped"]}, status=status.HTTP_400_BAD_REQUEST)

        workspace = Workspace.objects.get(slug=slug)
        strategy = request.data.get("duplicate_strategy", "skip")
        dedupe_by = request.data.get("dedupe_by", "auto")
        relations_mode = request.data.get("relations_mode", "merge")
        if relations_mode not in ("merge", "replace"):
            return Response({"relations_mode": ["Use merge or replace"]}, status=status.HTTP_400_BAD_REQUEST)
        # Off by default: a file row only matches PRE-EXISTING records, so rows
        # that share an identifier within the file don't collapse into one.
        dedupe_within_file = str(request.data.get("dedupe_within_file", "false")).lower() == "true"
        dry_run = str(request.data.get("dry_run", "false")).lower() == "true"
        invalid_row_strategy = request.data.get("invalid_row_strategy", "abort")
        try:
            row_overrides = json.loads(request.data.get("row_overrides", "{}"))
            value_overrides = json.loads(request.data.get("value_overrides", "{}"))
        except json.JSONDecodeError:
            return Response({"row_overrides": ["Must be valid JSON"]}, status=status.HTTP_400_BAD_REQUEST)
        if invalid_row_strategy not in ("abort", "skip"):
            return Response(
                {"invalid_row_strategy": ["Use abort or skip"]},
                status=status.HTTP_400_BAD_REQUEST,
            )
        result = {
            "total": len(rows),
            "created": 0,
            "updated": 0,
            "skipped": 0,
            "errors": [],
            "invalid_row_strategy": invalid_row_strategy,
            "aborted": False,
        }
        unparseable = {}
        touched = []
        run_created_ids = set()

        with transaction.atomic():
            for index, row in enumerate(rows, start=header_row + 1):
                effective_row, effective_mapping = _apply_row_overrides(
                    row, mapping, row_overrides.get(str(index), {})
                )
                effective_row, skip_row = _apply_overrides(effective_row, effective_mapping, value_overrides)
                if skip_row:
                    result["skipped"] += 1
                    continue
                if dry_run:
                    for field, token in _collect_unparsed(effective_row, effective_mapping).items():
                        tokens = unparseable.setdefault(field, {})
                        tokens[token] = tokens.get(token, 0) + 1
                try:
                    # Keep every row in its own savepoint. A database or validation
                    # error must not leave the remaining spreadsheet transaction aborted.
                    with transaction.atomic():
                        outcome, track = self._import_row(
                            workspace, effective_row, effective_mapping, strategy, defaults, dedupe_by,
                            relations_mode=relations_mode,
                            exclude_ids=None if dedupe_within_file else run_created_ids,
                        )
                    result[outcome] += 1
                    if outcome == "created" and track is not None:
                        run_created_ids.add(track.id)
                    if track is not None:
                        touched.append((track.id, outcome, index))
                except Exception as exc:
                    detail = _import_error_detail(index, exc, effective_row, effective_mapping)
                    result["errors"].append(detail)
                    if detail["code"] == "DATABASE_NOT_READY":
                        break
            aborted = result["errors"] and invalid_row_strategy == "abort"
            if not dry_run and not aborted:
                # Provenance: persist the source file itself + run + per-track
                # links. Re-importing the SAME file (stored asset, or same name
                # and size) reuses the asset so the filter shows one entry per file.
                upload.seek(0)
                content = upload.read()
                file_asset = source_asset or FileAsset.objects.filter(
                    workspace=workspace,
                    entity_type=FileAsset.EntityTypeContext.MUSIC_CATALOG,
                    attributes__music_asset_kind="IMPORT_SOURCE",
                    attributes__name=upload.name,
                    size=len(content),
                    is_deleted=False,
                ).first()
                if file_asset is None:
                    from plane.settings.storage import S3Storage

                    asset_key = f"{workspace.id}/music-imports/{timezone.now().strftime('%Y%m%d%H%M%S')}-{upload.name}"
                    storage = S3Storage()
                    storage.s3_client.put_object(
                        Bucket=storage.aws_storage_bucket_name, Key=asset_key, Body=content
                    )
                    file_asset = FileAsset.objects.create(
                        workspace=workspace,
                        entity_type=FileAsset.EntityTypeContext.MUSIC_CATALOG,
                        attributes={
                            "name": upload.name,
                            "type": getattr(upload, "content_type", "") or "text/csv",
                            "size": len(content),
                            "music_asset_kind": "IMPORT_SOURCE",
                            "upload_source": "manual",
                        },
                        asset=asset_key,
                        size=len(content),
                        is_uploaded=True,
                        created_by=request.user,
                    )
                run = _record_import_run(
                    workspace,
                    file_asset=file_asset,
                    source_name=upload.name,
                    source="MANUAL",
                    sheet=request.data.get("sheet"),
                    # The FULL panel configuration, so a re-import can restore
                    # it and the user only touches what changed.
                    rules={
                        "mapping": mapping,
                        "duplicate_strategy": strategy,
                        "dedupe_by": dedupe_by,
                        "relations_mode": relations_mode,
                        "dedupe_within_file": dedupe_within_file,
                        "value_overrides": value_overrides,
                        "row_overrides": row_overrides,
                        "invalid_row_strategy": invalid_row_strategy,
                        "defaults": defaults,
                    },
                    summary={k: result[k] for k in ("total", "created", "updated", "skipped")},
                    touched=touched,
                )
                result["import_run_id"] = str(run.id)
            if dry_run or aborted:
                transaction.set_rollback(True)
                result["aborted"] = bool(aborted)
        if dry_run:
            result["unparseable"] = {
                field: [{"value": token, "count": count} for token, count in tokens.items()]
                for field, tokens in unparseable.items()
            }
        return Response(result, status=status.HTTP_200_OK if not result["errors"] else status.HTTP_207_MULTI_STATUS)

    def _import_row(
        self, workspace, row, mapping, strategy, defaults=None, dedupe_by="auto", relations_mode="merge",
        exclude_ids=None,
    ):
        title = _mapped(row, mapping, "track.title")
        if not title:
            raise ValueError("Track title is empty")
        isrc = _mapped(row, mapping, "track.isrc").upper().replace("-", "")
        # Dedupe against SONGS only: music-video children share the parent's
        # title, so an unfiltered title match could wrongly hit the video.
        # `dedupe_by` lets the user pick WHICH identifier defines a duplicate
        # ("none" = always create, even with identical titles).
        # `exclude_ids` holds records CREATED earlier in this same import: when
        # de-duplication within the file is off, two file rows sharing a title
        # must not collapse — each becomes its own record instead of the second
        # one "matching" the first row's freshly-created track.
        songs = MusicTrack.objects.filter(workspace=workspace, parent_track__isnull=True)
        if exclude_ids:
            songs = songs.exclude(id__in=exclude_ids)
        track = None
        if dedupe_by == "none":
            track = None
        elif dedupe_by == "isrc":
            track = songs.filter(isrc__iexact=isrc).first() if isrc else None
        elif dedupe_by == "title":
            track = songs.filter(title__iexact=title).first()
        elif dedupe_by == "catalog":
            catalog = _mapped(row, mapping, "track.catalog")
            track = songs.filter(catalog__iexact=catalog).first() if catalog else None
        elif dedupe_by == "upc":
            upc = _mapped(row, mapping, "track.upc")
            track = songs.filter(upc__iexact=upc).first() if upc else None
        else:  # auto: ISRC first, then title + original release date
            track = songs.filter(isrc__iexact=isrc).first() if isrc else None
            if not track:
                track = songs.filter(
                    title__iexact=title,
                    original_release_date=_mapped_date(row, mapping, "track.original_release_date"),
                ).first()
        if track and strategy == "skip":
            return "skipped", track
        if track and strategy == "error":
            raise ValueError(f"Duplicate track: {isrc or title}")

        values = {
            "title": title,
            "version": _mapped(row, mapping, "track.version"),
            "isrc": isrc,
            "isrc_video": _mapped(row, mapping, "track.isrc_video").upper().replace("-", ""),
            "catalog": _mapped(row, mapping, "track.catalog"),
            "upc": _mapped(row, mapping, "track.upc"),
            "kind": _choice(_mapped(row, mapping, "track.kind"), MusicTrack.Kind.choices, MusicTrack.Kind.AUDIO),
            "status": _choice(_mapped(row, mapping, "track.status"), MusicTrack.Status.choices, MusicTrack.Status.DRAFT),
            "release_date": _mapped_date(row, mapping, "track.release_date"),
            "original_release_date": _mapped_date(row, mapping, "track.original_release_date"),
            "duration_ms": _duration_ms(_mapped(row, mapping, "track.duration_ms")),
            "country_of_recording": _mapped(row, mapping, "track.country_of_recording"),
            "language": _mapped(row, mapping, "track.language"),
            "explicit": _bool(_mapped(row, mapping, "track.explicit")),
            "ownership": _mapped(row, mapping, "track.ownership"),
            "us_publishing_obligations": _mapped(row, mapping, "track.us_publishing_obligations"),
            "recoupment": _mapped(row, mapping, "track.recoupment"),
            "digital_format": _mapped(row, mapping, "track.digital_format"),
            "lyrics": _mapped(row, mapping, "track.lyrics"),
            "aggregator_percentage": _decimal(_mapped(row, mapping, "track.aggregator_percentage")),
            "distributor_percentage": _decimal(_mapped(row, mapping, "track.distributor_percentage")),
            "record_label_percentage": _decimal(_mapped(row, mapping, "track.record_label_percentage")),
            "artist_percentage": _decimal(_mapped(row, mapping, "track.artist_percentage")),
            "writer_percentage": _decimal(_mapped(row, mapping, "track.writer_percentage")),
        }
        if track:
            for key, value in values.items():
                if f"track.{key}" in mapping:
                    setattr(track, key, value)
            track.save()
            outcome = "updated"
        else:
            track = MusicTrack.objects.create(workspace=workspace, **values)
            outcome = "created"

        release_title = _mapped(row, mapping, "release.title")
        if release_title:
            upc = _mapped(row, mapping, "release.upc")
            release = MusicRelease.objects.filter(workspace=workspace, upc__iexact=upc).first() if upc else None
            release = release or MusicRelease.objects.filter(
                workspace=workspace,
                title__iexact=release_title,
                release_date=_mapped_date(row, mapping, "release.release_date"),
            ).first()
            if not release:
                release = MusicRelease.objects.create(
                    workspace=workspace,
                    title=release_title,
                    release_type=_mapped(row, mapping, "release.type") or MusicRelease.Type.ALBUM,
                    upc=upc,
                    catalog_number=_mapped(row, mapping, "release.catalog_number"),
                    release_date=_mapped_date(row, mapping, "release.release_date"),
                    p_line=_mapped(row, mapping, "release.p_line"),
                    c_line=_mapped(row, mapping, "release.c_line"),
                )
            _restore_or_create(MusicReleaseTrack, workspace, {"release": release, "track": track})
        else:
            release = None

        role_fields = {
            "artists": (MusicParty.Kind.ARTIST, MusicCredit.Role.PRIMARY_ARTIST),
            "featured_artists": (MusicParty.Kind.ARTIST, MusicCredit.Role.FEATURED_ARTIST),
            "writers": (MusicParty.Kind.PERSON, MusicCredit.Role.WRITER),
            "authors": (MusicParty.Kind.PERSON, MusicCredit.Role.AUTHOR),
            "composers": (MusicParty.Kind.PERSON, MusicCredit.Role.COMPOSER),
            "producers": (MusicParty.Kind.PERSON, MusicCredit.Role.PRODUCER),
            "recording_engineers": (MusicParty.Kind.PERSON, MusicCredit.Role.RECORDING_ENGINEER),
            "mixers": (MusicParty.Kind.PERSON, MusicCredit.Role.MIXER),
            "mastering_engineers": (MusicParty.Kind.PERSON, MusicCredit.Role.MASTERING_ENGINEER),
            "legal_representatives": (MusicParty.Kind.PERSON, MusicCredit.Role.LEGAL_REPRESENTATIVE),
        }
        # People/genre/company fields accept SEVERAL mapped columns (e.g. two
        # writer columns); every column's value is split and imported.
        # relations_mode decides what happens with relations the track already
        # has: "merge" only ADDS what the file brings (nothing is removed);
        # "replace" makes mapped fields authoritative — relations no longer
        # coming from any mapped column are removed. Empty cells and unmapped
        # fields never touch existing relations in either mode.
        replace_relations = relations_mode == "replace"
        for field, (kind, role) in role_fields.items():
            produced_parties = set()
            for value in _mapped_many(row, mapping, field):
                for name in _split(value):
                    party = _party(workspace, name, kind)
                    produced_parties.add(party.id)
                    _restore_or_create(MusicCredit, workspace, {"track": track, "party": party, "role": role})
                    if release and role == MusicCredit.Role.PRIMARY_ARTIST:
                        _restore_or_create(
                            MusicReleaseArtist,
                            workspace,
                            {"release": release, "party": party, "role": MusicReleaseArtist.Role.PRIMARY},
                        )
            if replace_relations and produced_parties:
                _prune(track.credits.filter(role=role).exclude(party_id__in=produced_parties))
        produced_genres = set()
        for value in _mapped_many(row, mapping, "genres"):
            for name in _split(value):
                genre = _genre(workspace, name)
                produced_genres.add(genre.id)
                _restore_or_create(MusicTrackGenre, workspace, {"track": track, "genre": genre})
        if replace_relations and produced_genres:
            _prune(track.genre_links.exclude(genre_id__in=produced_genres))
        company_fields = {
            "record_label": MusicCompany.Kind.RECORD_LABEL,
            "aggregator": MusicCompany.Kind.AGGREGATOR,
            "distributor": MusicCompany.Kind.DISTRIBUTOR,
        }
        for field, kind in company_fields.items():
            produced_companies = set()
            for value in _mapped_many(row, mapping, field):
                for name in _split(value):
                    company = MusicCompany.objects.filter(workspace=workspace, kind=kind, name__iexact=name).first()
                    company = company or MusicCompany.objects.create(workspace=workspace, kind=kind, name=name)
                    produced_companies.add(company.id)
                    MusicDistribution.objects.get_or_create(workspace=workspace, track=track, company=company)
            if replace_relations and produced_companies:
                _prune(track.distributions.filter(company__kind=kind).exclude(company_id__in=produced_companies))
        self._import_links(workspace, track, row, mapping, values, replace_relations)
        self._apply_defaults(workspace, track, release, defaults or {})
        return outcome, track

    @staticmethod
    def _import_links(workspace, track, row, mapping, values, replace_relations=False):
        """Content-detected URL columns: a music-video URL materializes the
        video child track (with its ISRC/date when mapped) and both kinds
        attach as MusicLink rows — idempotent per URL so re-imports don't
        duplicate.
        """
        # Several link columns may map to the same field: the first video URL
        # materializes the video child, the rest attach as extra links on it.
        video_urls = [url for url in _mapped_many(row, mapping, "track.video_url") if url.startswith(("http://", "https://"))]
        if video_urls:
            video = track.videos.filter(kind=MusicTrack.Kind.MUSIC_VIDEO).first()
            video_release_date = _mapped_date(row, mapping, "track.video_release_date")
            if video is None:
                video = MusicTrack.objects.create(
                    workspace=workspace,
                    parent_track=track,
                    title=track.title,
                    kind=MusicTrack.Kind.MUSIC_VIDEO,
                    isrc_video=values.get("isrc_video", ""),
                    release_date=video_release_date,
                )
            else:
                changed = False
                if video_release_date and not video.release_date:
                    video.release_date = video_release_date
                    changed = True
                if values.get("isrc_video") and not video.isrc_video:
                    video.isrc_video = values["isrc_video"]
                    changed = True
                if changed:
                    video.save()
            for video_url in video_urls:
                if not video.links.filter(url=video_url).exists():
                    MusicLink.objects.create(
                        workspace=workspace,
                        track=video,
                        kind=MusicLink.Kind.MUSIC_VIDEO,
                        platform=_link_platform(video_url),
                        name=_link_platform(video_url) or "Music video",
                        url=video_url,
                        isrc=values.get("isrc_video", ""),
                    )
            # In replace mode mapped link columns are authoritative for their
            # kind: video links whose URL no longer comes from any mapped
            # column are removed
            if replace_relations:
                _prune(video.links.filter(kind=MusicLink.Kind.MUSIC_VIDEO).exclude(url__in=video_urls))

        streaming_urls = [
            url for url in _mapped_many(row, mapping, "track.streaming_url") if url.startswith(("http://", "https://"))
        ]
        for streaming_url in streaming_urls:
            if not track.links.filter(url=streaming_url).exists():
                MusicLink.objects.create(
                    workspace=workspace,
                    track=track,
                    kind=MusicLink.Kind.STREAMING,
                    platform=_link_platform(streaming_url),
                    name=_link_platform(streaming_url) or "Streaming",
                    url=streaming_url,
                    isrc=values.get("isrc", ""),
                )
        if replace_relations and streaming_urls:
            _prune(track.links.filter(kind=MusicLink.Kind.STREAMING).exclude(url__in=streaming_urls))

    @staticmethod
    def _apply_defaults(workspace, track, release, defaults):
        for entry in defaults.get("credit_entries", []):
            party = MusicParty.objects.get(id=entry["party_id"], workspace=workspace)
            credit = _restore_or_create(
                MusicCredit,
                workspace,
                {
                    "track": track,
                    "party": party,
                    "role": entry.get("role") or MusicCredit.Role.PRIMARY_ARTIST,
                },
            )
            for field in ("percentage", "publishing_share", "territory", "notes"):
                if field in entry:
                    setattr(credit, field, entry[field] or ("" if field in ("territory", "notes") else None))
            credit.save()
            if release and credit.role == MusicCredit.Role.PRIMARY_ARTIST:
                _restore_or_create(
                    MusicReleaseArtist,
                    workspace,
                    {"release": release, "party": party, "role": MusicReleaseArtist.Role.PRIMARY},
                )

        for genre_id in defaults.get("genre_ids", []):
            genre = MusicGenre.objects.get(id=genre_id, workspace=workspace)
            _restore_or_create(MusicTrackGenre, workspace, {"track": track, "genre": genre})

        for entry in defaults.get("distribution_entries", []):
            company = MusicCompany.objects.get(id=entry["company_id"], workspace=workspace)
            distribution, _ = MusicDistribution.objects.get_or_create(
                workspace=workspace,
                track=track,
                company=company,
            )
            for field in ("percentage", "territory", "valid_from", "valid_to"):
                if field in entry:
                    value = entry[field]
                    setattr(distribution, field, value or ("" if field == "territory" else None))
            distribution.save()

        for entry in defaults.get("releases", []):
            selected_release = MusicRelease.objects.get(id=entry["id"], workspace=workspace)
            _restore_or_create(
                MusicReleaseTrack,
                workspace,
                {"track": track, "release": selected_release},
                {
                    "disc_number": entry.get("disc_number", 1),
                    "track_number": entry.get("track_number", 1),
                },
            )


class MusicReportEndpoint(MusicBaseView):
    content_negotiation_class = MusicReportContentNegotiation

    @allow_permission([ROLE.ADMIN], level="WORKSPACE")
    def get(self, request, slug):
        today = date.today()
        window = request.query_params.get("window")
        params = request.query_params.copy()
        if window == "upcoming":
            params["from"] = today.isoformat()
        elif window == "last_30_days":
            params["from"] = (today - timedelta(days=30)).isoformat()
            params["to"] = today.isoformat()
        tracks = _filter_tracks(
            MusicTrack.objects.filter(workspace__slug=slug, parent_track__isnull=True).prefetch_related(
                "credits__party", "release_links__release", "genre_links__genre", "videos__links",
                "distributions__company",
            ),
            params,
        )
        headers = [
            "Song", "Version", "Artists", "Featured artists", "ISRC", "Status", "Release date",
            "Releases", "UPC", "Genres", "Videos", "Video ISRCs", "Video dates", "Video URLs",
            "Credits and rights", "Companies and shares", "Ownership", "P line",
        ]
        rows = []
        for track in tracks:
            releases = [link.release for link in track.release_links.all()]
            videos = list(track.videos.all())
            rows.append([
                track.title,
                track.version,
                "; ".join(c.party.display_name for c in track.credits.all() if c.role == MusicCredit.Role.PRIMARY_ARTIST),
                "; ".join(c.party.display_name for c in track.credits.all() if c.role == MusicCredit.Role.FEATURED_ARTIST),
                track.isrc,
                track.status,
                track.release_date or "",
                "; ".join(release.title for release in releases),
                "; ".join(filter(None, (release.upc for release in releases))),
                "; ".join(link.genre.name for link in track.genre_links.all()),
                "; ".join(video.title for video in videos),
                "; ".join(filter(None, (video.isrc_video for video in videos))),
                "; ".join(str(video.release_date or "") for video in videos),
                "; ".join(link.url for video in videos for link in video.links.all() if link.kind == MusicLink.Kind.MUSIC_VIDEO),
                "; ".join(
                    f'{credit.party.display_name} ({credit.get_role_display()}: {credit.percentage if credit.percentage is not None else "-"}%, publishing {credit.publishing_share if credit.publishing_share is not None else "-"}%)'
                    for credit in track.credits.all()
                ),
                "; ".join(
                    f'{distribution.company.name} ({distribution.company.get_kind_display()}: {distribution.percentage if distribution.percentage is not None else "-"}%)'
                    for distribution in track.distributions.all()
                ),
                track.ownership,
                track.p_line,
            ])
        export_format = request.query_params.get("format", "csv").lower()
        filename = f'music-catalog-{window or "report"}'
        if export_format == "xlsx":
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Music catalog"
            sheet.append(headers)
            for row in rows:
                sheet.append(row)
            output = BytesIO()
            workbook.save(output)
            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}.xlsx"'
            return response
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="{filename}.csv"'
        response.write("\ufeff")
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        return response
